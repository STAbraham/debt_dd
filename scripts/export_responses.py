#!/usr/bin/env python3
"""Export TRACKER.md to a dated outbound response spreadsheet.

Usage: python3 scripts/export_responses.py [YYYYMMDD]
Writes "Diligence Question Responses/Diligence Responses_<date>.xlsx".

Internal-only lines (**Note (internal):** ... and **Delivered:** ...) are
stripped; everything else in the A: block is exported verbatim. Items still
marked Drafted trigger a warning — review them before sending.
"""
import re
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SECTION_RE = re.compile(r"^## (\d+)\. (.+)$")
ITEM_RE = re.compile(r"^### (\S+) (.+) \[(\w+)\]$")
INTERNAL_FIELDS = ("**Note (internal):**", "**Delivered:**")


def parse(md: str):
    items, section = [], ""
    cur, field = None, None
    for line in md.splitlines():
        m = SECTION_RE.match(line)
        if m:
            section = m.group(2).strip()
            continue
        m = ITEM_RE.match(line)
        if m:
            cur = {"id": m.group(1), "topic": m.group(2).strip(), "status": m.group(3),
                   "section": section, "q": [], "a": []}
            items.append(cur)
            field = None
            continue
        if cur is None:
            continue
        if line.startswith("**Q:**"):
            field = "q"
            cur["q"].append(line[len("**Q:**"):].strip())
        elif line.startswith("**A:**"):
            field = "a"
            cur["a"].append(line[len("**A:**"):].strip())
        elif any(line.startswith(f) for f in INTERNAL_FIELDS):
            field = None
        elif field:
            cur[field].append(line)
    for it in items:
        for k in ("q", "a"):
            it[k] = "\n".join(it[k]).strip().replace("\\*", "*")
    return items


def main():
    stamp = sys.argv[1] if len(sys.argv) > 1 else date.today().strftime("%Y%m%d")
    items = parse((ROOT / "TRACKER.md").read_text())
    if not items:
        sys.exit("No items parsed from TRACKER.md — check heading format '### <id> <topic> [<Status>]'.")
    drafted = [it["id"] for it in items if it["status"] == "Drafted"]
    if drafted:
        print(f"WARNING: still Drafted (review before sending): {', '.join(drafted)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Diligence Tracker"
    ws.append(["Diligence Questions"])
    ws["A1"].font = Font(bold=True, size=14)
    headers = ["ID", "Section", "Topic", "Diligence question / request", "Status", "Response / evidence"]
    ws.append(headers)
    for c in ws[2]:
        c.font = Font(bold=True)
    for it in items:
        ws.append([it["id"], it["section"], it["topic"], it["q"], it["status"], it["a"]])
    widths = [8, 30, 28, 60, 10, 80]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=3):
        for c in row:
            c.alignment = wrap

    out = ROOT / "Diligence Question Responses" / f"Diligence Responses_{stamp}.xlsx"
    wb.save(out)
    print(f"Wrote {out.relative_to(ROOT)} ({len(items)} items)")


if __name__ == "__main__":
    main()
