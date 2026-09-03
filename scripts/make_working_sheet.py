#!/usr/bin/env python3
"""Generate the staging spreadsheet Steve edits: WORKING - Diligence Responses.xlsx.

Usage: python3 scripts/make_working_sheet.py [--force]

"Working" tab = every non-Closed item, with editable Response / Verdict /
"Notes to Claude" columns (yellow). "Closed (reference)" tab = what's already
been sent. Regenerated from TRACKER.md after every sync — don't hoard edits
in an old copy; edit, then tell Claude to "sync the working sheet".

Guard: refuses to overwrite the existing sheet if it contains edits that are
not reflected in TRACKER.md (Response text differs, or any Verdict/Notes
present), or if the workbook is currently open in Excel (~$ lock file).
Sync those edits first, or pass --force to overwrite deliberately.
"""
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from export_responses import parse

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "WORKING - Diligence Responses.xlsx"

EDIT_FILL = PatternFill("solid", fgColor="FFF7DE")
HDR_FILL = PatternFill("solid", fgColor="1F2937")
HDR_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")


def style_sheet(ws, widths, editable_cols=()):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.fill, c.font = HDR_FILL, HDR_FONT
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
            if c.column in editable_cols and ws.title == "Working":
                c.fill = EDIT_FILL
    ws.freeze_panes = "A2"


def unsynced_edits(items):
    """Return descriptions of edits in the existing sheet not reflected in TRACKER.md."""
    if not OUT.exists():
        return []
    from openpyxl import load_workbook
    tracker = {it["id"]: it for it in items}
    wb = load_workbook(OUT)
    if "Working" not in wb.sheetnames:
        return []
    found = []
    for row in wb["Working"].iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        vals = [(c if c is not None else "") for c in row[:7]]
        iid, _topic, _status, _q, resp, verdict, notes = (list(vals) + [""] * 7)[:7]
        t = tracker.get(str(iid))
        ta = t["a"] if t else ""
        if str(resp).replace("_x000B_", "\n").strip() != ta.strip():
            found.append(f"{iid}: Response text differs from tracker")
        if str(verdict).strip():
            found.append(f"{iid}: Verdict = {verdict}")
        if str(notes).strip():
            found.append(f"{iid}: Notes to Claude present")
    return found


def main():
    force = "--force" in sys.argv
    items = parse((ROOT / "TRACKER.md").read_text())

    lock = OUT.parent / f"~${OUT.name}"
    if lock.exists() and not force:
        sys.exit(f"REFUSING to overwrite: {OUT.name} is open in Excel (lock file {lock.name}). "
                 "Close it (saving first), sync any edits, then rerun — or pass --force.")
    edits = unsynced_edits(items)
    if edits and not force:
        sys.exit("REFUSING to overwrite: the existing sheet has edits not yet in TRACKER.md:\n  "
                 + "\n  ".join(edits)
                 + "\nSync the working sheet first, or pass --force to discard them.")
    open_items = [it for it in items if it["status"] != "Closed"]
    closed = [it for it in items if it["status"] == "Closed"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Working"
    ws.append(["ID", "Topic", "Status", "Question", "Response (edit me — your text wins)",
               "Verdict", "Notes to Claude", "Claude's internal context (won't be sent)"])
    for it in open_items:
        ws.append([it["id"], it["topic"], it["status"], it["q"], it["a"], "", "", it["note"]])
    dv = DataValidation(type="list", formula1='"Approved,Claude: revise (see notes),Discuss"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"F2:F{max(len(open_items) + 1, 2)}")
    style_sheet(ws, [7, 26, 10, 45, 70, 24, 40, 45], editable_cols=(5, 6, 7))

    ref = wb.create_sheet("Closed (reference)")
    ref.append(["ID", "Topic", "Question", "Response as sent", "Delivered"])
    for it in closed:
        ref.append([it["id"], it["topic"], it["q"], it["a"], it["delivered"]])
    style_sheet(ref, [7, 26, 45, 70, 30])

    wb.save(OUT)
    print(f"Wrote {OUT.name}: {len(open_items)} working items, {len(closed)} closed for reference")


if __name__ == "__main__":
    main()
